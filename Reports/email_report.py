"""
Shared logic for sending daily attendance report emails.
Used by the management command and the API view.
"""

import io
from datetime import date
from pathlib import Path

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from Organization.models import Office
from Reports.utils import (
    fetch_attendance_report_rows_for_office,
    get_recipients_for_office,
)


def build_excel_bytes(rows, office_name, report_date_str):
    """Build Excel file (parent-only columns) and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    headers = [
        "Employee Code",
        "Employee Name",
        "Device ID",
        "Log Date",
        "First In",
        "Last Out",
        "Hours Worked",
        "Punches",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row.get("employee_code") or "")
        ws.cell(row=row_idx, column=2, value=row.get("employee_name") or "")
        ws.cell(row=row_idx, column=3, value=row.get("device_id") or "")
        ws.cell(row=row_idx, column=4, value=row.get("log_date") or "")
        ws.cell(row=row_idx, column=5, value=row.get("first_in") or "")
        ws.cell(row=row_idx, column=6, value=row.get("last_out") or "")
        ws.cell(row=row_idx, column=7, value=row.get("hours_worked") or "")
        ws.cell(row=row_idx, column=8, value=row.get("punches") or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _find_logo_path():
    """Find attenova-logo in project root (png, jpg, jpeg)."""
    base = Path(settings.BASE_DIR)
    for ext in (".png", ".jpg", ".jpeg"):
        p = base / f"attenova-logo{ext}"
        if p.exists():
            return p
    return None


def send_report_email(
    recipients,
    office,
    report_date,
    rows,
    dry_run=False,
    to_emails_override: list | None = None,
    *,
    attachment_label: str | None = None,
    report_period_label: str | None = None,
    matrix_attachment: tuple[bytes, str] | None = None,
):
    """Render HTML, attach flat Excel (cron) or matrix CSV (manual range), send email."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage
    from email.mime.base import MIMEBase
    from email import encoders

    office_name = attachment_label if attachment_label else office.name
    if report_period_label:
        report_date_str = report_period_label
    else:
        report_date_str = report_date.strftime("%b %d, %Y")
    year = report_date.year

    hero_image_path = Path(settings.BASE_DIR) / "Biometric-image.jpg"
    has_hero_image = hero_image_path.exists()

    logo_path = _find_logo_path()
    has_logo = logo_path is not None

    html_body = render_to_string(
        "Reports/daily_attendance_email.html",
        {
            "office_name": office_name,
            "report_date": report_date_str,
            "year": year,
            "has_hero_image": has_hero_image,
            "has_logo": has_logo,
        },
    )

    if matrix_attachment:
        excel_bytes, filename = matrix_attachment
    else:
        excel_bytes = build_excel_bytes(rows, office_name, report_date_str)
        filename = (
            f"Daily_Attendance_{office_name.replace(' ', '_')}_"
            f"{report_date.strftime('%Y-%m-%d')}.xlsx"
        )

    if matrix_attachment:
        subject = f"[Attenova] Attendance Matrix — {office_name} ({report_date_str})"
    else:
        subject = f"[Attenova] Daily Attendance Report — {office_name} ({report_date_str})"
    if to_emails_override:
        to_emails = list(dict.fromkeys(e for e in to_emails_override if e))
    else:
        to_emails = list({u.email for u in recipients if u.email} | {"akashyadav181198@gmail.com"})
    if not to_emails:
        return

    if dry_run:
        return

    # Top-level: mixed (so we have one body part + one attachment)
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = settings.DEFAULT_FROM_EMAIL
    msg["To"] = ", ".join(to_emails)
    msg["Reply-To"] = settings.DEFAULT_FROM_EMAIL

    # Body: multipart/related so HTML + inline images are one "body" (images won’t show as attachments)
    html_part = MIMEText(html_body, "html")
    html_part.add_header("Content-ID", "<html_root>")
    body_related = MIMEMultipart("related")
    body_related.set_param("start", "<html_root>")
    body_related.attach(html_part)

    if has_hero_image:
        with open(hero_image_path, "rb") as f:
            img = MIMEImage(f.read(), _subtype="jpeg")
            img.add_header("Content-ID", "<biometric_hero>")
            img.add_header("Content-Disposition", "inline")
            body_related.attach(img)

    if logo_path:
        ext = logo_path.suffix.lower()
        subtype = "png" if ext == ".png" else "jpeg"
        with open(logo_path, "rb") as f:
            logo_img = MIMEImage(f.read(), _subtype=subtype)
            logo_img.add_header("Content-ID", "<attenova_logo>")
            logo_img.add_header("Content-Disposition", "inline")
            body_related.attach(logo_img)

    msg.attach(body_related)

    if matrix_attachment:
        excel_part = MIMEBase("text", "csv")
        excel_part.set_payload(excel_bytes)
        encoders.encode_base64(excel_part)
        excel_part.add_header(
            "Content-Disposition",
            "attachment",
            filename=filename,
        )
    else:
        excel_part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        excel_part.set_payload(excel_bytes)
        encoders.encode_base64(excel_part)
        excel_part.add_header(
            "Content-Disposition",
            "attachment",
            filename=filename,
        )
    msg.attach(excel_part)

    # SMTP backend uses message.as_bytes(linesep="\r\n") which Python 3.13's MIME doesn't support
    if settings.EMAIL_BACKEND == "django.core.mail.backends.smtp.EmailBackend":
        use_tls = getattr(settings, "EMAIL_USE_TLS", True)
        use_ssl = getattr(settings, "EMAIL_USE_SSL", False)
        port = getattr(settings, "EMAIL_PORT", 587)
        host = getattr(settings, "EMAIL_HOST", "smtp.gmail.com")
        user = getattr(settings, "EMAIL_HOST_USER", "") or None
        password = getattr(settings, "EMAIL_HOST_PASSWORD", "") or None
        smtp_cls = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
        with smtp_cls(host, port) as smtp:
            if use_tls and not use_ssl:
                smtp.starttls()
            if user and password:
                smtp.login(user, password)
            smtp.sendmail(settings.DEFAULT_FROM_EMAIL, to_emails, msg.as_bytes())
    else:
        # Console / other backends use as_bytes() without linesep
        from django.core.mail import get_connection

        class _RawMessage:
            def __init__(self, mime_msg, from_email, to_list, subject):
                self._msg = mime_msg
                self.from_email = from_email
                self.to = to_list
                self.subject = subject
                self.encoding = None

            def message(self):
                return self._msg

            def recipients(self):
                return self.to

        connection = get_connection()
        raw = _RawMessage(msg, settings.DEFAULT_FROM_EMAIL, to_emails, subject)
        connection.send_messages([raw])


def run_send_daily_attendance_emails(
    report_date: date,
    office_id_filter: int | None = None,
    dry_run: bool = False,
):
    """
    Send daily attendance report emails for all offices (or a single office).
    Cronjob mode: sends to each office's Admin, Manager, Supervisor + akashyadav181198@gmail.com.
    Returns {"sent": int, "message": str}.
    """
    offices = Office.objects.filter(is_active=True).select_related("organization")
    if office_id_filter:
        offices = offices.filter(pk=office_id_filter)
    offices = list(offices)

    sent = 0
    for office in offices:
        recipients = get_recipients_for_office(office)
        if not recipients:
            continue
        rows = fetch_attendance_report_rows_for_office(office.id, report_date)
        send_report_email(recipients, office, report_date, rows, dry_run)
        sent += 1

    return {
        "sent": sent,
        "message": f"Report sent to {sent} office(s).",
    }


def _format_report_period(start_date: date, end_date: date) -> str:
    if start_date == end_date:
        return start_date.strftime("%b %d, %Y")
    if start_date.year == end_date.year:
        return f"{start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}"
    return f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"


def run_send_manual_ui_report(
    user,
    start_date: date,
    end_date: date,
    office_id: int | None = None,
    search: str = "",
):
    """
    Manual UI: send matrix CSV for [start_date, end_date] to user (+ configured copy).
    """
    from Reports.matrix_export import build_matrix_csv_bundle_for_user

    csv_bytes, filename, office, attachment_label = build_matrix_csv_bundle_for_user(
        user=user,
        start_date=start_date,
        end_date=end_date,
        office_id=office_id,
        search=search or "",
    )
    period = _format_report_period(start_date, end_date)
    to_emails = list(
        dict.fromkeys(
            [e for e in [user.email, "akashyadav181198@gmail.com"] if e]
        )
    )
    send_report_email(
        [],
        office,
        end_date,
        [],
        dry_run=False,
        to_emails_override=to_emails,
        attachment_label=attachment_label,
        report_period_label=period,
        matrix_attachment=(csv_bytes, filename),
    )
    return {
        "sent": 1,
        "message": f"Attendance matrix sent for {period}.",
    }
